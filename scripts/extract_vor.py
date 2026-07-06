#!/usr/bin/env python3
"""
extract_vor.py — Извлечение Ведомости объёмов работ (ВОР) из OCR-результатов.

Поддерживает:
- Буронабивные сваи (объём бетона, бурения, выбурки грунта, арматура)
- Автоматический расчёт по формулам
- Выход: CSV, TXT, Excel (.xlsx)

Использование:
    # Для свай
    python extract_vor.py --ocr-json ./output/piles_qwen.json --type piles --output ./output/vor_piles.xlsx
    
    # Для геологии  
    python extract_vor.py --ocr-json ./output/kj7_02_qwen.json --type geology --output ./output/vor_geo.csv
"""

import argparse
import csv
import json
import math
import re
import sys
from pathlib import Path


def load_ocr_data(json_path: Path) -> list:
    """Load OCR JSON results."""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def extract_pile_specifications(data: list) -> list:
    """Extract pile specifications from OCR data with flexible parsing."""
    specs = []
    
    all_text = ' '.join(
        ' '.join(r.get('text_lines', []))
        for r in data
    )
    
    # Try multiple patterns for pile marks and quantities
    # Pattern 1: Mark followed by quantity (flexible spacing)
    pile_patterns = [
        # СБн12-450 100
        re.compile(r'([СC][БB][нnN]\d{1,2}-\d{3,4})\s+(\d+)', re.IGNORECASE),
        # Сваи буронабивные СБн12-450 (100 шт)
        re.compile(r'буронабивные\s+([СC][БB][нnN]\d{1,2}-\d{3,4})\s*.*?\((\d+)\s*шт', re.IGNORECASE),
        # Ведомость свай: СБн12-450 - 100 шт
        re.compile(r'([СC][БB][нnN]\d{1,2}-\d{3,4})\s*[-–—]\s*(\d+)\s*шт', re.IGNORECASE),
        # Flexible: СБн12-450 followed by number within 50 chars (for tables)
        re.compile(r'([СC][БB][нnN]\d{1,2}-\d{3,4})\s*[^0-9]{0,50}(\d+)', re.IGNORECASE),
    ]
    
    seen = set()
    for pattern in pile_patterns:
        matches = pattern.findall(all_text)
        for mark, count_str in matches:
            mark = mark.upper().replace('N', 'Н').replace('B', 'Б').replace('C', 'С')
            count = int(count_str)
            
            if mark in seen:
                continue
            seen.add(mark)
            
            # Parse dimensions
            dim_match = re.match(r'СБН(\d+)-(\d+)', mark)
            if dim_match:
                length_m = int(dim_match.group(1))
                diameter_mm = int(dim_match.group(2))
            else:
                length_m = 12
                diameter_mm = 450
            
            # Calculate volumes
            radius_m = (diameter_mm / 1000) / 2
            volume_per_pile = math.pi * (radius_m ** 2) * length_m
            total_concrete = volume_per_pile * count
            total_soil = total_concrete * 1.15
            total_rebar_kg = total_concrete * 120  # 120 kg/m³
            
            specs.append({
                'mark': mark,
                'diameter_mm': diameter_mm,
                'length_m': length_m,
                'count': count,
                'volume_per_pile_m3': round(volume_per_pile, 3),
                'total_concrete_m3': round(total_concrete, 2),
                'total_burrowing_m3': round(total_concrete, 2),
                'total_soil_extraction_m3': round(total_soil, 2),
                'total_rebar_kg': round(total_rebar_kg, 1),
                'total_rebar_t': round(total_rebar_kg / 1000, 3),
            })
    
    # If no quantities found, just extract marks with default count=1
    if not specs:
        mark_pattern = re.compile(r'[СC][БB][нnN]\d{1,2}-\d{3,4}', re.IGNORECASE)
        marks = mark_pattern.findall(all_text)
        for mark in marks:
            mark = mark.upper().replace('N', 'Н').replace('B', 'Б').replace('C', 'С')
            if mark in seen:
                continue
            seen.add(mark)
            
            dim_match = re.match(r'СБН(\d+)-(\d+)', mark)
            if dim_match:
                length_m = int(dim_match.group(1))
                diameter_mm = int(dim_match.group(2))
            else:
                length_m = 12
                diameter_mm = 450
            
            radius_m = (diameter_mm / 1000) / 2
            volume_per_pile = math.pi * (radius_m ** 2) * length_m
            
            specs.append({
                'mark': mark,
                'diameter_mm': diameter_mm,
                'length_m': length_m,
                'count': 1,  # Unknown quantity
                'volume_per_pile_m3': round(volume_per_pile, 3),
                'total_concrete_m3': round(volume_per_pile, 2),
                'total_burrowing_m3': round(volume_per_pile, 2),
                'total_soil_extraction_m3': round(volume_per_pile * 1.15, 2),
                'total_rebar_kg': round(volume_per_pile * 120, 1),
                'total_rebar_t': round(volume_per_pile * 120 / 1000, 3),
            })
    
    return specs


def generate_pile_vor(specs: list) -> list:
    """Generate ВОР (Ведомость объёмов работ) for piles."""
    rows = []
    
    # Title
    rows.append({'№ п/п': '', 'Код': '', 'Наименование': 'ВЕДОМОСТЬ ОБЪЁМОВ РАБОТ', 'Ед.изм': '', 'Количество': '', 'Объем': ''})
    rows.append({'№ п/п': '', 'Код': '', 'Наименование': 'Буронабивные сваи', 'Ед.изм': '', 'Количество': '', 'Объем': ''})
    
    row_num = 1
    totals = {'concrete': 0, 'burrowing': 0, 'soil': 0, 'rebar': 0, 'count': 0}
    
    for spec in specs:
        # Drilling
        rows.append({
            '№ п/п': row_num,
            'Код': 'Е2-196',
            'Наименование': f"Бурение скважин D={spec['diameter_mm']}мм, L={spec['length_m']}м ({spec['mark']})",
            'Ед.изм': 'м³',
            'Количество': spec['count'],
            'Объем': spec['total_burrowing_m3'],
        })
        totals['burrowing'] += spec['total_burrowing_m3']
        row_num += 1
        
        # Rebar cage
        rows.append({
            '№ п/п': row_num,
            'Код': 'Е4-48',
            'Наименование': f"Установка арматурных каркасов свай {spec['mark']}",
            'Ед.изм': 'шт',
            'Количество': spec['count'],
            'Объем': f"~{spec['total_rebar_t']} т",
        })
        totals['rebar'] += spec['total_rebar_t']
        row_num += 1
        
        # Concrete
        rows.append({
            '№ п/п': row_num,
            'Код': 'Е4-1',
            'Наименование': f"Бетонирование свай {spec['mark']}",
            'Ед.изм': 'м³',
            'Количество': spec['count'],
            'Объем': spec['total_concrete_m3'],
        })
        totals['concrete'] += spec['total_concrete_m3']
        row_num += 1
        
        # Soil extraction
        rows.append({
            '№ п/п': row_num,
            'Код': 'Е2-120',
            'Наименование': f"Выбурка грунта ({spec['mark']})",
            'Ед.изм': 'м³',
            'Количество': spec['count'],
            'Объем': spec['total_soil_extraction_m3'],
        })
        totals['soil'] += spec['total_soil_extraction_m3']
        row_num += 1
        
        totals['count'] += spec['count']
    
    # Total
    rows.append({
        '№ п/п': '',
        'Код': '',
        'Наименование': 'ИТОГО:',
        'Ед.изм': '',
        'Количество': totals['count'],
        'Объем': f"Бетон: {round(totals['concrete'], 2)} м³ | Арм.: {round(totals['rebar'], 3)} т | Выбурка: {round(totals['soil'], 2)} м³",
    })
    
    return rows, totals


def save_vor_csv(rows: list, output_path: Path):
    """Save ВОР to CSV."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ['№ п/п', 'Код', 'Наименование', 'Ед.изм', 'Количество', 'Объем']
    
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
        writer.writeheader()
        writer.writerows(rows)
    
    print(f"Saved CSV: {output_path}")


def save_vor_excel(rows: list, totals: dict, output_path: Path):
    """Save ВОР to Excel with formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("WARNING: openpyxl not installed, skipping Excel output", file=sys.stderr)
        return
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "ВОР"
    
    # Styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'ВЕДОМОСТЬ ОБЪЁМОВ РАБОТ'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = 'Буронабивные сваи'
    ws['A2'].font = Font(bold=True, size=11, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['№ п/п', 'Код ресурса', 'Наименование работ и затрат', 'Ед.изм', 'Количество', 'Объем']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        cell.border = thin_border
    
    # Data
    row_idx = 5
    for row in rows[2:]:  # Skip title rows
        if not row['№ п/п'] and row['Наименование'] == 'ИТОГО:':
            # Total row
            for col in range(1, 7):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = thin_border
                if col == 1:
                    cell.value = 'ИТОГО:'
                    cell.font = Font(bold=True)
                elif col == 5:
                    cell.value = row['Количество']
                    cell.font = Font(bold=True)
                elif col == 6:
                    cell.value = row['Объем']
                    cell.font = Font(bold=True)
            row_idx += 1
        elif row['№ п/п']:
            for col, key in enumerate(['№ п/п', 'Код', 'Наименование', 'Ед.изм', 'Количество', 'Объем'], 1):
                cell = ws.cell(row=row_idx, column=col, value=row[key])
                cell.border = thin_border
                if col == 3:
                    cell.alignment = Alignment(wrap_text=True)
            row_idx += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25
    
    wb.save(output_path)
    print(f"Saved Excel: {output_path}")


def save_vor_txt(rows: list, output_path: Path):
    """Save ВОР to human-readable TXT."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("ВЕДОМОСТЬ ОБЪЁМОВ РАБОТ\n")
        f.write("Буронабивные сваи\n")
        f.write("=" * 80 + "\n\n")
        
        for row in rows:
            if row['№ п/п'] == '':
                if row['Наименование'] == 'ИТОГО:':
                    f.write(f"\n{'='*80}\n")
                    f.write(f"ИТОГО: {row['Объем']}\n")
                    f.write(f"{'='*80}\n")
                else:
                    f.write(f"\n{row['Наименование']}\n")
            else:
                f.write(f"\n{row['№ п/п']}. [{row['Код']}] {row['Наименование']}\n")
                f.write(f"   Ед.изм: {row['Ед.изм']}, Кол-во: {row['Количество']}, Объем: {row['Объем']}\n")
    
    print(f"Saved TXT: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Extract ВОР from OCR results")
    parser.add_argument("--ocr-json", required=True, help="Path to OCR JSON")
    parser.add_argument("--type", default="piles", choices=["piles", "concrete", "geology"], help="Work type")
    parser.add_argument("--output", default="./output/vor.xlsx", help="Output path (.csv or .xlsx)")
    args = parser.parse_args()
    
    ocr_path = Path(args.ocr_json)
    if not ocr_path.exists():
        print(f"ERROR: File not found: {ocr_path}", file=sys.stderr)
        sys.exit(1)
    
    print(f"Loading OCR: {ocr_path}")
    data = load_ocr_data(ocr_path)
    print(f"Loaded {len(data)} records")
    
    output_path = Path(args.output)
    
    if args.type == "piles":
        specs = extract_pile_specifications(data)
        
        if not specs:
            print("ERROR: No pile specs found. Check OCR contains 'СБн12-450' etc.", file=sys.stderr)
            sys.exit(1)
        
        print(f"\nFound {len(specs)} pile types:")
        for spec in specs:
            print(f"  {spec['mark']}: {spec['count']} шт. (D={spec['diameter_mm']}мм, L={spec['length_m']}м)")
            print(f"    Бетон: {spec['total_concrete_m3']} м³ | Арм.: {spec['total_rebar_t']} т | Выбурка: {spec['total_soil_extraction_m3']} м³")
        
        total_count = sum(s['count'] for s in specs)
        print(f"\nВСЕГО свай: {total_count}")
        
        rows, totals = generate_pile_vor(specs)
        
        # Save all formats
        save_vor_csv(rows, output_path.with_suffix('.csv'))
        save_vor_excel(rows, totals, output_path.with_suffix('.xlsx'))
        save_vor_txt(rows, output_path.with_suffix('.txt'))
        
        print(f"\n{'='*60}")
        print("✅ ВОР СФОРМИРОВАНА")
        print(f"{'='*60}")
        print(f"Сваи: {totals['count']} шт.")
        print(f"Бетон: {round(totals['concrete'], 2)} м³")
        print(f"Арматура: {round(totals['rebar'], 3)} т")
        print(f"Выбурка грунта: {round(totals['soil'], 2)} м³")
        print(f"{'='*60}")
        
    else:
        print(f"Type '{args.type}' not yet implemented", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
