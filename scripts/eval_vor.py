#!/usr/bin/env python3
"""
Eval VOR (Ведомость объёмов работ) — проверка качества ВОР из чертежей.

Сравнивает сгенерированную ВОР с эталонной структурой от профессионала.
Проверяет: структуру, состав позиций, формулы, коэффициенты, пропущенные элементы.

Usage:
    python3 eval_vor.py --vor /path/to/ВОР.xlsx --example data/vor_structure_example.json
    python3 eval_vor.py --vor /path/to/ВОР.xlsx --mode full

Exit codes:
    0 — PASS (все проверки пройдены)
    1 — FAIL (есть критические ошибки)
    2 — WARNING (есть замечания, но не критичные)
"""

import argparse
import json
import sys
import openpyxl
from pathlib import Path


class VOREvaluator:
    """Проверка ВОР на соответствие профессиональному стандарту."""

    def __init__(self, example_path="data/vor_structure_example.json"):
        self.example = json.load(open(example_path, 'r', encoding='utf-8'))
        self.checks = []
        self.failures = []
        self.warnings = []

    def log(self, status, message):
        self.checks.append((status, message))
        if status == "FAIL":
            self.failures.append(message)
        elif status == "WARN":
            self.warnings.append(message)
        print(f"  [{status}] {message}")

    def check_structure(self, vor_items):
        """Проверка структуры разделов."""
        print("\n=== 1. СТРУКТУРА РАЗДЕЛОВ ===")

        sections = {}
        for item in vor_items:
            sec = item.get('section', 'БЕЗ_РАЗДЕЛА')
            if sec not in sections:
                sections[sec] = []
            sections[sec].append(item)

        expected_sections = list(self.example['structure'].keys())
        actual_sections = list(sections.keys())

        # Проверяем наличие всех разделов
        for sec in expected_sections:
            if sec in actual_sections:
                self.log("PASS", f"Раздел '{sec}' присутствует ({len(sections[sec])} позиций)")
            else:
                self.log("FAIL", f"Раздел '{sec}' ОТСУТСТВУЕТ! Ожидалось: {expected_sections}")

        # Проверяем лишние разделы
        for sec in actual_sections:
            if sec not in expected_sections and sec != 'БЕЗ_РАЗДЕЛА':
                self.log("WARN", f"Лишний раздел '{sec}' ({len(sections[sec])} позиций)")

        # Проверяем количество позиций по разделам
        for sec_name, sec_data in self.example['structure'].items():
            if sec_name in sections:
                actual_count = len(sections[sec_name])
                expected_count = sec_data['items_count']
                if actual_count < expected_count * 0.5:
                    self.log("FAIL", f"Раздел '{sec_name}': {actual_count} позиций, ожидалось ~{expected_count} (меньше 50%)")
                elif actual_count < expected_count * 0.8:
                    self.log("WARN", f"Раздел '{sec_name}': {actual_count} позиций, ожидалось ~{expected_count} (меньше 80%)")
                else:
                    self.log("PASS", f"Раздел '{sec_name}': {actual_count} позиций (ожидалось ~{expected_count})")

    def check_subsections(self, vor_items):
        """Проверка подразделов (фундаменты)."""
        print("\n=== 2. ПОДРАЗДЕЛЫ ФУНДАМЕНТОВ ===")

        fund_items = [i for i in vor_items if 'Фундамент' in i.get('section', '')]
        subsections = {}
        for item in fund_items:
            sub = item.get('subsection', 'БЕЗ_ПОДРАЗДЕЛА')
            if sub not in subsections:
                subsections[sub] = []
            subsections[sub].append(item)

        expected_subs = self.example['structure']['Раздел 2. Фундаменты']['subsections']

        for sub_name, sub_data in expected_subs.items():
            if sub_name in subsections:
                actual_count = len(subsections[sub_name])
                expected_count = sub_data['items_count']
                self.log("PASS", f"Подраздел '{sub_name}': {actual_count} позиций (ожидалось ~{expected_count})")
            else:
                self.log("FAIL", f"Подраздел '{sub_name}' ОТСУТСТВУЕТ! Критично для полноты ВОР")

    def check_required_elements(self, vor_items):
        """Проверка обязательных элементов."""
        print("\n=== 3. ОБЯЗАТЕЛЬНЫЕ ЭЛЕМЕНТЫ ===")

        all_names = ' '.join([i.get('name', '').lower() for i in vor_items])

        required_elements = [
            ("Временное ограждение", ["временн", "ограждение", "забор"]),
            ("Разработка грунта", ["разработка", "грунт", "котлован"]),
            ("Перевозка грунта", ["перевозка", "груз", "самосвал"]),
            ("ПГС подготовка", ["пгс", "песчано-гравий", "основание"]),
            ("Бетонная подготовка", ["подготовка", "бетон"]),
            ("Бетон В7.5", ["в7.5", "м100"]),
            ("Бетон В25", ["в25", "м350"]),
            ("Арматура каркас", ["каркас", "кп1"]),
            ("Арматура детали", ["детали", "а500", "а400"]),
            ("Опалубка", ["щит", "опалубка", "доска"]),
            ("Закладные детали", ["закладн", "мн"]),
            ("Анкерные блоки", ["анкер", "аб"]),
            ("Гидроизоляция", ["гидроизоляция", "герметик"]),
            ("Деформационные швы", ["деформацион", "шов"]),
            ("Утеплитель", ["пеноплекс", "утеплитель", "пенопол"]),
            ("Покраска металла", ["покрас", "грунтовка"]),
            ("Отмостка", ["отмостка"]),
        ]

        for element_name, keywords in required_elements:
            found = any(kw in all_names for kw in keywords)
            if found:
                self.log("PASS", f"Элемент '{element_name}' найден")
            else:
                self.log("FAIL", f"Элемент '{element_name}' НЕ НАЙДЕН! Проверить чертёж")

    def check_formulas(self, vor_items):
        """Проверка формул расчёта."""
        print("\n=== 4. ФОРМУЛЫ РАСЧЁТА ===")

        formulas = [i.get('formula', '') for i in vor_items if i.get('formula')]

        # Проверяем наличие коэффициентов уплотнения
        compaction_found = any('1.18' in f or '1,18' in f for f in formulas)
        if compaction_found:
            self.log("PASS", "Коэффициент уплотнения ПГС (1.18) найден")
        else:
            self.log("WARN", "Коэффициент уплотнения ПГС (1.18) НЕ НАЙДЕН! Для ПГС обязателен: 1.18 × 1.01")

        # Проверяем перевод единиц
        unit_1000 = any('/ 1000' in f or '/1000' in f for f in formulas)
        unit_100 = any('/ 100' in f or '/100' in f for f in formulas)

        if unit_1000:
            self.log("PASS", "Перевод в 1000 м³ найден (для земляных работ)")
        else:
            self.log("WARN", "Перевод в 1000 м³ не найден — проверить единицы земляных работ")

        if unit_100:
            self.log("PASS", "Перевод в 100 м³ найден (для бетона/засыпки)")
        else:
            self.log("WARN", "Перевод в 100 м³ не найден — проверить единицы бетона")

        # Проверяем массу арматуры
        rebar_formula = any('1000' in f and ('*' in f or '/') in f for f in formulas if 'арм' in str(f).lower() or 'детал' in str(f).lower())
        # Проверяем по именам позиций
        rebar_items = [i for i in vor_items if 'арм' in i.get('name', '').lower() or 'а500' in i.get('name', '').lower() or 'а400' in i.get('name', '').lower()]
        rebar_with_formula = [i for i in rebar_items if i.get('formula')]
        if rebar_with_formula:
            self.log("PASS", f"Арматура с формулами: {len(rebar_with_formula)} позиций")
        else:
            self.log("WARN", "Арматура без формул! Должно быть: количество × масса п.м. / 1000")

    def check_units(self, vor_items):
        """Проверка единиц измерения."""
        print("\n=== 5. ЕДИНИЦЫ ИЗМЕРЕНИЯ ===")

        units = {}
        for item in vor_items:
            u = item.get('unit', '')
            units[u] = units.get(u, 0) + 1

        expected_units = {
            '1000 м3': 'земляные работы',
            '100 м3': 'бетон, засыпка',
            'м3': 'материалы, бетон',
            'т': 'арматура, металл',
            'м2': 'опалубка, гидроизоляция',
            'м п.': 'швы, изоляция',
            'шт': 'элементы, ограждения'
        }

        for unit, desc in expected_units.items():
            if unit in units:
                self.log("PASS", f"Единица '{unit}' ({desc}): {units[unit]} позиций")
            else:
                self.log("WARN", f"Единица '{unit}' ({desc}) не найдена — проверить, нужна ли")

    def check_compaction(self, vor_items):
        """Проверка коэффициентов уплотнения."""
        print("\n=== 6. КОЭФФИЦИЕНТЫ УПЛОТНЕНИЯ ===")

        pgs_items = [i for i in vor_items if 'пгс' in i.get('name', '').lower() or 'песчано-гравий' in i.get('name', '').lower()]

        if not pgs_items:
            self.log("WARN", "ПГС не найден в ВОР — проверить чертёж на наличие песчано-гравийной подготовки")
            return

        for item in pgs_items:
            formula = item.get('formula', '')
            name = item.get('name', '')
            if '1.18' in formula or '1,18' in formula:
                self.log("PASS", f"ПГС '{name[:50]}': коэфф. уплотнения 1.18 найден")
            elif '1.01' in formula or '1,01' in formula:
                self.log("PASS", f"ПГС '{name[:50]}': коэфф. 1.01 найден")
            else:
                self.log("FAIL", f"ПГС '{name[:50]}': коэфф. уплотнения НЕ НАЙДЕН! Должно быть: объем × 1.18 × 1.01")

    def check_rebar_calculation(self, vor_items):
        """Проверка расчёта арматуры."""
        print("\n=== 7. РАСЧЁТ АРМАТУРЫ ===")

        rebar_items = [i for i in vor_items if 'арм' in i.get('name', '').lower() or 'а500' in i.get('name', '').lower() or 'а400' in i.get('name', '').lower()]

        if not rebar_items:
            self.log("FAIL", "Арматура не найдена в ВОР! Критично — проверить чертёж")
            return

        total_rebar = sum(float(i.get('qty', 0) or 0) for i in rebar_items)
        self.log("PASS", f"Арматура: {len(rebar_items)} позиций, общая масса ~{total_rebar:.2f} т")

        # Проверяем формулы
        items_with_formula = [i for i in rebar_items if i.get('formula')]
        if len(items_with_formula) < len(rebar_items) * 0.5:
            self.log("WARN", f"Арматура без формул: {len(rebar_items) - len(items_with_formula)} позиций. Должно быть: кол-во × масса п.м. / 1000")
        else:
            self.log("PASS", f"Арматура с формулами: {len(items_with_formula)} / {len(rebar_items)} позиций")

    def check_completeness(self, vor_items):
        """Общая проверка полноты."""
        print("\n=== 8. ПОЛНОТА ВОР ===")

        total_items = len(vor_items)
        self.log("INFO", f"Всего позиций в ВОР: {total_items}")

        if total_items < 50:
            self.log("FAIL", f"Слишком мало позиций: {total_items}. Ожидалось ~134 (по эталону)")
        elif total_items < 100:
            self.log("WARN", f"Мало позиций: {total_items}. Возможно, пропущены вспомогательные работы или элементы")
        else:
            self.log("PASS", f"Количество позиций: {total_items} (в норме)")

        # Проверяем наличие комментариев/формул
        items_with_formula = [i for i in vor_items if i.get('formula')]
        if len(items_with_formula) < total_items * 0.3:
            self.log("WARN", f"Мало формул: {len(items_with_formula)} / {total_items}. Формулы помогают проверить расчёт")
        else:
            self.log("PASS", f"Формул: {len(items_with_formula)} / {total_items}")

    def evaluate(self, vor_items):
        """Запуск всех проверок."""
        print("=" * 60)
        print("EVAL VOR — Проверка Ведомости объёмов работ")
        print("=" * 60)
        print(f"Эталон: {self.example['meta']['object']}")
        print(f"Версия: {self.example['meta']['version']}")
        print()

        self.check_structure(vor_items)
        self.check_subsections(vor_items)
        self.check_required_elements(vor_items)
        self.check_formulas(vor_items)
        self.check_units(vor_items)
        self.check_compaction(vor_items)
        self.check_rebar_calculation(vor_items)
        self.check_completeness(vor_items)

        # Результат
        print("\n" + "=" * 60)
        print("РЕЗУЛЬТАТ")
        print("=" * 60)

        total = len(self.checks)
        passed = len([c for c in self.checks if c[0] == "PASS"])
        failed = len(self.failures)
        warnings = len(self.warnings)

        print(f"  Проверок: {total}")
        print(f"  PASS: {passed}")
        print(f"  FAIL: {failed}")
        print(f"  WARN: {warnings}")

        if failed > 0:
            print(f"\n  ❌ FAIL: {failed} критических ошибок. ВОР требует доработки!")
            print("  Критические ошибки:")
            for f in self.failures:
                print(f"    - {f}")
            return 1
        elif warnings > 3:
            print(f"\n  ⚠️ WARNING: {warnings} замечаний. Рекомендуется проверить перед отправкой.")
            return 2
        else:
            print(f"\n  ✅ PASS: ВОР прошла проверку. Можно передавать в smeta.")
            return 0


def load_vor_from_excel(path):
    """Загрузка ВОР из Excel."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    items = []
    current_section = 'БЕЗ_РАЗДЕЛА'
    current_subsection = None

    for i in range(1, ws.max_row + 1):
        row = ws[i]
        a = row[0].value
        b = row[1].value
        c = row[2].value
        d = row[3].value
        e = row[4].value
        f = row[5].value
        g = row[6].value

        # Раздел
        if isinstance(a, str) and 'Раздел' in a:
            current_section = a
            continue

        # Подраздел
        if a and isinstance(a, str) and not a.startswith('=') and c is None and not a.isdigit():
            current_subsection = a
            continue

        # Позиция
        if b and c and isinstance(c, str) and not c.startswith('='):
            items.append({
                'section': current_section,
                'subsection': current_subsection,
                'n': len(items) + 1,
                'lsr': b,
                'name': c.strip(),
                'unit': d,
                'qty': e,
                'refs': str(f).strip() if f else '',
                'formula': str(g).strip() if g else ''
            })

    return items


def main():
    parser = argparse.ArgumentParser(description='Eval VOR — проверка Ведомости объёмов работ')
    parser.add_argument('--vor', required=True, help='Путь к ВОР.xlsx')
    parser.add_argument('--example', default='data/vor_structure_example.json', help='Путь к эталонной структуре')
    args = parser.parse_args()

    evaluator = VOREvaluator(args.example)
    vor_items = load_vor_from_excel(args.vor)

    exit_code = evaluator.evaluate(vor_items)
    sys.exit(exit_code)


if __name__ == '__main__':
    main()
