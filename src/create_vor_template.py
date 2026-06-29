import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ВОР"

# Заголовок
ws['A1'] = 'ВЕДОМОСТЬ ОБЪЁМОВ РАБОТ (ВОР)'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:F1')

ws['A2'] = 'Объект:'
ws['B2'] = '[Указать наименование объекта]'
ws['A3'] = 'Масштаб:'
ws['B3'] = '[Указать масштаб, например 1:100]'
ws['A3'].font = Font(bold=True, color="FF0000")
ws['B3'].font = Font(color="FF0000")

ws['A4'] = 'Дата:'
ws['B4'] = '[Дата составления]'

ws['A5'] = 'Составил:'
ws['B5'] = '[ФИО]'

# Шапка таблицы
headers = ['№ п/п', 'Наименование работ', 'Ед. изм.', 'Количество', 'Шифр/марка', 'Примечание', 'Источник (лист чертежа)']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=7, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

# Пример строки (для参考)
example = ['1', 'Установка вентиляторов радиальных [марка]', 'шт', '1', 'ВР 80-75 №6.3', 'периметром до 1600 мм', 'ОВ-1, лист 3']
for col, val in enumerate(example, 1):
    cell = ws.cell(row=8, column=col, value=val)
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

# Форматирование
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 25
ws.column_dimensions['G'].width = 20

ws.row_dimensions[7].height = 30

# Примечание
ws['A10'] = 'ПРИМЕЧАНИЯ:'
ws['A10'].font = Font(bold=True)
ws['A11'] = '• Масштаб обязателен для расчёта объёмов с планов/разрезов'
ws['A12'] = '• Колонка "Источник" — обязательна для трассировки (на каком листе найдено)'
ws['A13'] = '• Неочевидные позиции помечать "УТОЧНИТЬ"'
ws['A14'] = '• Спецификации оборудования — копировать как есть, проверить соответствие чертежам'

wb.save('/root/.openclaw/workspace/skills/drawings-to-vor/templates/ВОР_шаблон.xlsx')
print("Шаблон ВОР создан: templates/ВОР_шаблон.xlsx")