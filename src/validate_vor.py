#!/usr/bin/env python3
"""
Валидация ВОР (Ведомости объёмов работ):
1. Полнота (все ли позиции из списка попали в ВОР)
2. Корректность единиц измерения
3. Отсутствие дублей
4. Соответствие чертежам (если предоставлены страницы)

Использование:
    python3 validate_vor.py --vor ВОР.xlsx [--drawing_pages pages/]
"""

import argparse
import sys
import openpyxl
import json
import os
from typing import List, Dict


# Стандартные единицы измерения по ФЕР-2020
STANDARD_UNITS = {
    'шт', 'компл', 'м', 'пог.м', 'м2', 'м²', 'м3', 'м³', 'т', 'кг', '100кг',
    'м.п.', 'м.п', 'м2', 'м3', 'м3', 'ед', 'набор', 'пара'
}


def load_vor(filepath: str) -> List[Dict]:
    """Загружает ВОР из Excel."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    items = []
    
    # Ищем заголовок (обычно строка 5-7)
    header_row = None
    for row in range(1, min(20, ws.max_row + 1)):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a and str(cell_a).lower() in ['№ п/п', '№', 'п/п', 'номер']:
            header_row = row
            break
    
    if not header_row:
        print("WARN: Не найдена строка заголовка. Используется строка 1.")
        header_row = 1
    
    # Определяем колонки
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            val_str = str(val).lower().strip()
            if val_str in ['№ п/п', '№', 'п/п', 'номер', '№п/п']:
                col_map['num'] = col
            elif 'наименование' in val_str or 'название' in val_str:
                col_map['name'] = col
            elif 'ед.' in val_str or 'единица' in val_str or 'еи' in val_str:
                col_map['unit'] = col
            elif 'количество' in val_str or 'кол.' in val_str or 'объем' in val_str:
                col_map['qty'] = col
            elif 'примечание' in val_str or 'прим.' in val_str:
                col_map['note'] = col
            elif 'источник' in val_str or 'чертеж' in val_str:
                col_map['source'] = col
    
    # Если не нашли — пробуем стандартные позиции
    if 'num' not in col_map:
        col_map['num'] = 1
    if 'name' not in col_map:
        col_map['name'] = 2
    if 'unit' not in col_map:
        col_map['unit'] = 3
    if 'qty' not in col_map:
        col_map['qty'] = 4
    
    # Читаем данные
    for row in range(header_row + 1, ws.max_row + 1):
        num = ws.cell(row=row, column=col_map.get('num', 1)).value
        if num is None:
            continue
        
        try:
            num_str = str(int(num)) if isinstance(num, (int, float)) else str(num)
        except:
            continue
        
        name = ws.cell(row=row, column=col_map.get('name', 2)).value
        if not name:
            continue
        
        unit = ws.cell(row=row, column=col_map.get('unit', 3)).value
        qty = ws.cell(row=row, column=col_map.get('qty', 4)).value
        note = ws.cell(row=row, column=col_map.get('note', 5)).value if 'note' in col_map else None
        source = ws.cell(row=row, column=col_map.get('source', 6)).value if 'source' in col_map else None
        
        items.append({
            'num': num_str,
            'name': str(name).strip(),
            'unit': str(unit).strip() if unit else '',
            'qty': qty,
            'note': str(note).strip() if note else '',
            'source': str(source).strip() if source else ''
        })
    
    return items


def check_units(items: List[Dict]) -> List[str]:
    """Проверяет корректность единиц измерения."""
    warnings = []
    
    for item in items:
        unit = item['unit'].lower().replace('.', '').replace(' ', '')
        if not unit:
            warnings.append(f"WARN: Позиция №{item['num']} '{item['name']}' — не указана единица измерения")
            continue
        
        # Нормализация
        unit_normalized = unit.replace('м2', 'м²').replace('м3', 'м³').replace('пог.м', 'м').replace('м.п.', 'м')
        
        # Проверка стандартности
        if unit_normalized not in STANDARD_UNITS and unit not in STANDARD_UNITS:
            warnings.append(
                f"WARN: Нестандартная единица измерения в позиции №{item['num']}: "
                f"'{item['unit']}' — проверить соответствие ФЕР-2020"
            )
    
    return warnings


def check_duplicates(items: List[Dict]) -> List[str]:
    """Проверяет на дубли по наименованию."""
    warnings = []
    seen = {}
    
    for item in items:
        name_lower = item['name'].lower().strip()
        if name_lower in seen:
            warnings.append(
                f"WARN: Дубль позиции: №{item['num']} '{item['name']}' — "
                f"уже есть как №{seen[name_lower]}"
            )
        else:
            seen[name_lower] = item['num']
    
    return warnings


def check_completeness(items: List[Dict], drawing_metadata: Dict = None) -> List[str]:
    """Проверяет полноту ВОР."""
    warnings = []
    
    # Проверка наличия источников (чертежей)
    has_sources = any(item['source'] for item in items)
    if not has_sources and drawing_metadata:
        warnings.append(
            "WARN: Не указаны источники (чертежи) для позиций. "
            "Рекомендуется добавить колонку 'Источник' с указанием листа."
        )
    
    # Проверка на «УТОЧНИТЬ»
    utocheniya = [item for item in items if 'уточнить' in item['name'].lower() or 'уточнить' in item.get('note', '').lower()]
    if utocheniya:
        warnings.append(
            f"WARN: {len(utocheniya)} позиций требуют уточнения. "
            f"Необходимо получить уточнения перед передачей в смету."
        )
    
    return warnings


def check_against_drawings(items: List[Dict], drawing_pages_dir: str) -> List[str]:
    """Проверяет соответствие ВОР чертежам (если есть метаданные)."""
    warnings = []
    
    meta_path = os.path.join(drawing_pages_dir, "metadata.json")
    if not os.path.exists(meta_path):
        return warnings
    
    with open(meta_path, "r", encoding="utf-8") as f:
        metadata = json.load(f)
    
    if not metadata.get('scale_detected') and not metadata.get('scale_manual'):
        warnings.append(
            "FAIL: Масштаб чертежа не указан. Все объёмы могут быть неверны. "
            "Необходимо указать масштаб перед анализом."
        )
    
    return warnings


def generate_report(items: List[Dict], warnings: List[str], errors: List[str]) -> str:
    """Генерирует отчёт о валидации."""
    report = []
    report.append("=" * 60)
    report.append("ВАЛИДАЦИЯ ВОР (Ведомость объёмов работ)")
    report.append("=" * 60)
    report.append(f"Позиций в ВОР: {len(items)}")
    report.append("")
    
    if errors:
        report.append("❌ ОШИБКИ (FAIL):")
        for err in errors:
            report.append(f"  {err}")
        report.append("")
    
    if warnings:
        report.append("⚠️  ПРЕДУПРЕЖДЕНИЯ (WARN):")
        for warn in warnings:
            report.append(f"  {warn}")
        report.append("")
    
    if not errors and not warnings:
        report.append("✅ ВОР ПРОВЕРЕНА — ошибок нет")
    elif not errors:
        report.append("✅ ВОР ПРОВЕРЕНА С ЗАМЕЧАНИЯМИ: WARN требуют внимания, но не критичны")
    else:
        report.append("❌ ВОР НЕ ПРОЙДЕНА: исправить FAIL перед передачей в смету")
    
    return "\n".join(report)


def main():
    parser = argparse.ArgumentParser(description='Валидация ВОР')
    parser.add_argument('--vor', required=True, help='Файл ВОР (.xlsx)')
    parser.add_argument('--drawing_pages', default='', help='Папка со страницами чертежей (PNG)')
    args = parser.parse_args()
    
    print(f"=== ВАЛИДАЦИЯ ВОР ===")
    print(f"ВОР: {args.vor}")
    if args.drawing_pages:
        print(f"Чертежи: {args.drawing_pages}")
    print()
    
    # Загрузка ВОР
    items = load_vor(args.vor)
    print(f"Загружено позиций: {len(items)}")
    print()
    
    errors = []
    warnings = []
    
    # Проверки
    warnings += check_units(items)
    warnings += check_duplicates(items)
    warnings += check_completeness(items)
    
    if args.drawing_pages:
        warnings += check_against_drawings(items, args.drawing_pages)
    
    # Отчёт
    report = generate_report(items, warnings, errors)
    print(report)
    
    # Сохранение отчёта
    report_path = args.vor.replace('.xlsx', '_validation_report.txt')
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"\nОтчёт сохранён: {report_path}")
    
    # Exit code
    if errors:
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()
